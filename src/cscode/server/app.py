"""CScode Server — FastAPI application with Event Sourcing session architecture.

Phase 2 architecture:
- SessionV2 + EventStore: append-only event log, no delete+reinsert
- SessionCoordinator: per-session serialization
- AgentV2 per request via factory
- SessionProjector: rebuilds state from events
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import sys
import time
import traceback
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, AsyncGenerator, AsyncIterator

from fastapi import APIRouter, FastAPI, HTTPException, Request, UploadFile, WebSocket
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from cscode.app.factory import create_agent_v2, create_tool_registry
from cscode.core.compression import ContextCompressor
from cscode.core.coordinator import SessionCoordinator
from cscode.core.external_directory import ExternalDirectoryStore
from cscode.core.session import SessionLockManager, SessionProjector, SessionV2
from cscode.core.sharing import ShareStore
from cscode.core.tracker import TaskTracker
from cscode.core.workspace import WorkspaceStore
from cscode.llm.types import LLMRequest
from cscode.lsp.manager import LSPManager
from cscode.schema.events import (
    Error,
    Finish,
    LLMEvent,
    Pending,
    ReasoningDelta,
    ReasoningEnded,
    ReasoningStarted,
    TextDelta,
    TextEnded,
    TextStarted,
    ToolCallDelta,
    ToolCallEnded,
    ToolCallStarted,
    ToolFailure,
    ToolResult,
)
from cscode.schema.ids import SessionID
from cscode.schema.messages import (
    Message as NewMessage,
)
from cscode.schema.messages import (
    MessageRole,
)
from cscode.server.compactor import Compactor
from cscode.server.integration import IntegrationTokenStore, WebSocketManager
from cscode.server.projector import Projector
from cscode.server.question_registry import QuestionRegistry
from cscode.server.routes import permissions_router, sessions_router
from cscode.server.state import state
from cscode.storage.db import Database
from cscode.storage.event_store import EventStore
from cscode.tools2.pty import PTYInput, PTYSessionManager

logger = logging.getLogger(__name__)

api_router = APIRouter(prefix="/api")

OUTPUTS_DIR = Path("/tmp/cscode-outputs")

# Event types that are persisted to EventStore for message history.
# text.delta is intentionally excluded: streaming deltas are real-time only via SSE.
# text.ended (final complete text) is persisted for session history replay.
PERSIST_EVENT_TYPES = frozenset(
    {
        "step.started",
        "text.ended",
        "step.ended",
        "tool.called",
        "tool.success",
        "tool.failed",
        "error",
    }
)


async def _project_events(
    session_id: str,
    events: list[dict[str, object]],
) -> None:
    """Append events to EventStore and notify the Projector (messages table).

    Guards against uninitialized _event_store / _projector.
    """
    if _event_store is None:
        return
    appended = await _event_store.append(session_id, events)
    if _projector is not None:
        for evt in appended:
            await _projector.on_event(evt)


def _llm_event_to_dict(event: LLMEvent) -> dict[str, object]:
    """Convert LLMEvent to dict for SSE streaming.

    Returns dict with 'type' and 'data' keys where data holds the payload.
    This ensures frontend applyEvent receives event.data.xxx consistently.
    """
    match event:
        case TextStarted():
            return {"type": "step.started", "data": {}}
        case TextDelta(text=text):
            return {"type": "text.delta", "data": {"content": text}}
        case TextEnded(full_text=full_text):
            return {"type": "text.ended", "data": {"content": full_text}}
        case ToolCallStarted(tool_call_id=id, name=name):
            return {"type": "tool.called", "data": {"tool_call_id": id, "name": name, "args": {}}}
        case ToolCallDelta(tool_call_id=id, args_text=args_text):
            return {"type": "tool.call_delta", "data": {"tool_call_id": id, "args_text": args_text}}
        case ToolCallEnded(tool_call_id=id, name=name, args=args):
            return {"type": "tool.called", "data": {"tool_call_id": id, "name": name, "args": args}}
        case ToolResult(
            tool_call_id=id,
            result=result,
            tool_name=tool_name,
            tool_args=tool_args,
            metadata=metadata,
        ):
            return {
                "type": "tool.success",
                "data": {
                    "tool_call_id": id,
                    "result": result,
                    "name": tool_name,
                    "args": tool_args,
                    "metadata": metadata,
                },
            }
        case Finish(finish_reason=finish_reason):
            return {"type": "complete", "data": {"finish_reason": finish_reason}}
        case ToolFailure(
            tool_call_id=id,
            error=error,
            tool_name=tool_name,
            tool_args=tool_args,
            metadata=metadata,
        ):
            return {
                "type": "tool.failed",
                "data": {
                    "tool_call_id": id,
                    "error": error,
                    "name": tool_name,
                    "args": tool_args,
                    "metadata": metadata,
                },
            }
        case Error(error=error):
            # Include full structured error detail: [REASON] module.method — message
            if isinstance(error, Exception):
                detail = f"{type(error).__name__}: {error}"
            else:
                detail = str(error)
            return {"type": "error", "data": {"content": detail}}
        case Pending():
            return {"type": "status", "data": {"message": "pending"}}
        case ReasoningStarted():
            return {"type": "reasoning", "data": {"status": "started"}}
        case ReasoningDelta(text=text):
            return {"type": "reasoning", "data": {"delta": text}}
        case ReasoningEnded():
            return {"type": "reasoning", "data": {"status": "ended"}}
        case _:
            logger.warning("_llm_event_to_dict: unhandled event type=%s", type(event).__name__)
            return {"type": "unknown", "data": {}}


async def _auto_compact(session_id: str, event_store: EventStore) -> None:
    """Fire-and-forget event compaction via Compactor (runs in background)."""
    global _compactor
    if _compactor is None:
        return
    try:
        events = await event_store.read(session_id)
        if not events:
            return
        message_count = sum(
            1
            for e in events
            if e.type in ("prompt.admitted", "text.ended", "tool.success", "tool.failed")
        )
        if message_count < 20:
            logger.debug("_auto_compact: skip session=%s count=%d < 20", session_id, message_count)
            return
        await _compactor.compact(session_id)
    except Exception:
        logger.exception("auto_compact failed for %s", session_id)


def _build_system_prompt(file_context: str = "") -> NewMessage:
    base = (
        "You are CScode, an AI-powered coding assistant. You help users write, review, and debug code. "
        "You have access to tools for reading, writing, and editing files, "
        "running shell commands, searching codebases, and browsing the web."
    )
    if file_context:
        base += f"\n\n{file_context}"
    base += "\n\nCRITICAL RULES FOR TESTING — VIOLATION WILL BE DETECTED:\n"
    base += "1. Every test case MUST be executed through real tool calls (browser, bash, etc.).\n"
    base += "   Each tool call is recorded and verified. You CANNOT fake execution.\n"
    base += "2. NEVER infer or guess test results from documentation, code, or prior knowledge.\n"
    base += "   If you did not call a tool, the result does not exist.\n"
    base += "3. If a test cannot be executed (no credentials, blocked URL, timeout):\n"
    base += "   Mark it SKIPPED — do NOT mark it as passed or failed.\n"
    base += "4. For browser tests, you MUST capture BOTH screenshot AND HTML content.\n"
    base += "   A test without both is UNVERIFIED and will not count as executed.\n"
    base += "5. task_id format MUST be: TC-XXX (XXX is 3-digit number, e.g. TC-001, TC-002).\n"
    base += "6. In your final response, use this format for each test case:\n"
    base += "   [EXECUTED]   TC-001 — Login success — evidence: screenshot + HTML\n"
    base += "   [FAILED]     TC-002 — Login failure — error: timeout\n"
    base += "   [SKIPPED]    TC-003 — Payment test — reason: no test credentials\n"
    base += "   [UNVERIFIED] TC-004 — Empty page — re-run needed\n"
    base += "7. The verification report is generated from the database, not from your text.\n"
    base += "   You cannot convince the system — only real tool calls count."
    return NewMessage.system(base)


class _CallableProcessor:
    """Adapts a callable agent runner to core/coordinator's processor interface."""

    def __init__(self, handler: Any) -> None:
        self._handler = handler

    async def process(self, session_id: str) -> str:
        result = await self._handler()
        return str(result) if result else ""


@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncIterator[None]:
    global \
        _db, \
        _event_store, \
        _coordinator, \
        _projector, \
        _compactor, \
        _tracker, \
        _question_registry, \
        _tool_registry, \
        _workspace_store, \
        _external_dir_store, \
        _ws_manager, \
        _token_store, \
        _pty_manager, \
        _share_store

    from cscode.server.state import state as app_state

    # Diagnostics log — stderr + file
    from cscode.utils.logging import setup_logging

    setup_logging("DEBUG")

    fh = logging.FileHandler("/tmp/cscode-diag.log", mode="w")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
    logging.getLogger().addHandler(fh)
    logging.getLogger().setLevel(logging.DEBUG)
    logger.info("=== CScode server started (Event Sourcing architecture) ===")

    resource_dir = os.environ.get("CSCODE_RESOURCE_DIR", "")
    if resource_dir:
        python_dir = os.path.join(resource_dir, "python")
        if os.path.isdir(python_dir):
            existing = os.environ.get("PYTHONPATH", "")
            os.environ["PYTHONPATH"] = (
                f"{python_dir}{os.pathsep}{existing}" if existing else python_dir
            )
            logger.debug("Set PYTHONPATH for subprocesses: %s", python_dir)

    db_path = os.environ.get("CSCODE_DB_PATH")
    _db = Database(db_path=db_path)
    app_state.db = _db
    await _db.init()

    # New architecture: EventStore + SessionCoordinator
    _event_store = EventStore(_db)
    app_state.event_store = _event_store
    _coordinator = SessionCoordinator()
    app_state.coordinator = _coordinator
    _projector = Projector(_db)
    app_state.projector = _projector
    _compactor = Compactor(_db, _event_store, _projector)
    app_state.compactor = _compactor
    _tracker = TaskTracker(_db)
    app_state.tracker = _tracker
    _question_registry = QuestionRegistry()
    app_state.question_registry = _question_registry

    # Shared tool registry (AgentV2 instances will reuse this)
    _tool_registry = create_tool_registry()
    app_state.tool_registry = _tool_registry

    # Workspace store (P2-3)
    _workspace_store = WorkspaceStore(_db) if _db else None
    app_state.workspace_store = _workspace_store

    # Share store (P1-2)
    _share_store = ShareStore(_db) if _db else None
    app_state.share_store = _share_store

    # External directory registry (P2-16)
    _external_dir_store = ExternalDirectoryStore()
    app_state.external_dir_store = _external_dir_store

    _ws_manager = WebSocketManager(
        event_store=_event_store,
        chat_handler=_ws_chat_handler,
    )
    app_state.ws_manager = _ws_manager
    await _ws_manager.start_event_bridge()
    _token_store = IntegrationTokenStore()
    app_state.token_store = _token_store
    _pty_manager = PTYSessionManager()
    app_state.pty_manager = _pty_manager

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUTS_DIR / "evidence").mkdir(parents=True, exist_ok=True)
    template_path = OUTPUTS_DIR / "xlsx_template.py"
    if not template_path.exists():
        template_path.write_text(_XLSX_TEMPLATE)
        template_path.chmod(0o755)

    logger.info("Lifespan startup complete")

    yield

    if _ws_manager is not None:
        await _ws_manager.stop_event_bridge()
    if _db is not None:
        await _db.close()
    logger.info("Lifespan shutdown complete")


app = FastAPI(title="CScode API", version="0.3.4", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Route modules (split from app.py for maintainability)
app.include_router(sessions_router)
app.include_router(permissions_router)


@app.middleware("http")
async def log_requests(request: Request, call_next: Any) -> Response:
    """Log method, path, status, duration for every request."""
    start = time.monotonic()
    response: Response = await call_next(request)
    duration_ms = (time.monotonic() - start) * 1000
    logger.info(
        "%s %s %s %.0fms",
        request.method,
        request.url.path,
        response.status_code,
        duration_ms,
    )
    return response


_db: Database | None = None
_event_store: EventStore | None = None
_coordinator: SessionCoordinator | None = None
_projector: Projector | None = None
_compactor: Compactor | None = None
_tracker: TaskTracker | None = None
_question_registry: QuestionRegistry | None = None
_tool_registry: Any = None
_workspace_store: WorkspaceStore | None = None
_share_store: ShareStore | None = None
_external_dir_store: ExternalDirectoryStore | None = None
_active_agent_tasks: dict[str, asyncio.Task[Any]] = {}
_session_queues: dict[str, asyncio.Queue[dict[str, object]]] = {}


class ChatResponse(BaseModel):
    response: str
    session_id: str


class McpServerConfig(BaseModel):
    name: str
    command: str
    args: list[str] = []
    env: dict[str, str] = {}
    enabled: bool = True


class PluginConfig(BaseModel):
    enabled: list[str] = []
    settings: dict[str, dict[str, object]] = {}


class PermissionRule(BaseModel):
    pattern: str
    allow: bool = True


class ConfigRequest(BaseModel):
    provider: str = "openai"
    model: str = "gpt-4o"
    api_base: str | None = None
    api_key: str | None = None
    max_tokens: int = 4096
    temperature: float = 0.7
    top_p: float = 1.0
    system_prompt: str | None = None
    theme: str | None = None
    mcp_servers: list[McpServerConfig] = []
    plugins: PluginConfig = PluginConfig()
    permission_rules: list[PermissionRule] = []
    keybindings: dict[str, str] = {}


class SessionCreateRequest(BaseModel):
    title: str = "New Session"


@api_router.get("/health")
async def health() -> dict[str, str]:
    from cscode import __version__

    return {"status": "ok", "version": __version__}


async def _get_or_create_session(
    session_id: str | None,
    event_store: EventStore,
    config_data: dict[str, Any] | None = None,
) -> tuple[SessionV2, bool]:
    """Get existing session or create new one. Returns (session_v2, is_new)."""
    if session_id is None:
        session_id = str(uuid.uuid4())

    is_new = False
    sid = SessionID(session_id)
    events = await event_store.read(sid)
    if events:
        state = SessionProjector.project(events)
        session_v2 = SessionV2(event_store, sid, state)
    else:
        provider = "openai"
        model = "gpt-4o"
        if config_data:
            provider = config_data.get("provider", "openai")
            model = config_data.get("model", "gpt-4o")
        session_v2 = await SessionV2.create(
            event_store, model=model, provider=provider, title="New Chat", agent="auto"
        )
        is_new = True

    return session_v2, is_new


async def _build_context_messages(
    session_v2: SessionV2,
    user_message: str,
    file_context: str = "",
) -> list[NewMessage]:
    """Build message list for LLM from session state + new user input."""
    # Get messages from session state (reconstructed from events)
    messages = list(SessionProjector.build_context(session_v2.state))

    if not messages or messages[0].role != MessageRole.SYSTEM:
        # Merge file_context into system prompt (some providers ignore 2+ system messages)
        messages.insert(0, _build_system_prompt(file_context))
    elif file_context:
        messages.append(NewMessage.system(file_context))

    # Add user message
    messages.append(NewMessage.user(user_message.strip() if user_message else "请分析附件内容"))

    return messages


@api_router.post("/chat")
async def chat(request: Request) -> ChatResponse:
    message = ""
    session_id: str | None = None
    files: list[tuple[str, bytes]] = []

    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type:
        form = await request.form()
        _message = form.get("message", "")
        if isinstance(_message, str):
            message = _message
        _session_raw = form.get("session_id", None)
        if isinstance(_session_raw, str):
            session_id = _session_raw
        for f in form.getlist("files"):
            if isinstance(f, UploadFile):
                content = await f.read()
                files.append((f.filename or "unknown", content))
    else:
        import json as _json

        try:
            body = await request.json()
        except _json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON body")
        message = body.get("message", "")
        session_id = body.get("session_id", None)
        files_data = body.get("files", [])
        for f in files_data:
            if isinstance(f, dict):
                import base64

                try:
                    file_bytes = base64.b64decode(f.get("content", ""))
                except Exception:
                    file_bytes = f.get("content", "").encode("utf-8")
                files.append((f.get("name", "file"), file_bytes))

    return await _handle_chat(message, session_id, files)


@api_router.post("/chat/stream")
async def chat_stream(request: Request) -> StreamingResponse:
    message = ""
    session_id: str | None = None
    files: list[tuple[str, bytes]] = []

    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type or "application/x-www-form-urlencoded" in content_type:
        form = await request.form()
        _message = form.get("message", "")
        if isinstance(_message, str):
            message = _message
        _session_raw = form.get("session_id", None)
        if isinstance(_session_raw, str):
            session_id = _session_raw
        for f in form.getlist("files"):
            if isinstance(f, UploadFile):
                content = await f.read()
                files.append((f.filename or "unknown", content))
    else:
        body = await request.json()
        message = body.get("message", "")
        session_id = body.get("session_id", None)
        files_data = body.get("files", [])
        for f in files_data:
            if isinstance(f, dict):
                import base64

                try:
                    file_bytes = base64.b64decode(f.get("content", ""))
                except Exception:
                    file_bytes = f.get("content", "").encode("utf-8")
                files.append((f.get("name", "file"), file_bytes))

    async def event_stream() -> AsyncGenerator[str, None]:
        global _event_store, _coordinator, _tool_registry, _db
        nonlocal session_id
        if _event_store is None or _coordinator is None:
            yield f"data: {json.dumps({'type': 'error', 'content': 'Server not initialized'})}\n\n"
            return

        session_v2: Any = None
        agent_task: Any = None

        try:
            if session_id is None:
                session_id = str(uuid.uuid4())
            # Load or create session
            config_data = None
            if _db is not None:
                from cscode.core.config import ConfigStore

                store = ConfigStore(_db)
                saved_config_dict = await store.get()
                if saved_config_dict:
                    config_data = saved_config_dict

            session_v2, is_new = await _get_or_create_session(session_id, _event_store, config_data)

            # Per-session concurrency lock: reject if already processing
            if not await SessionLockManager.try_lock(str(session_v2.session_id)):
                yield f"data: {json.dumps({'type': 'error', 'content': 'Session is already processing'})}\n\n"
                return

            if is_new:
                yield f"data: {json.dumps({'type': 'session', 'session_id': session_v2.session_id})}\n\n"

            # Build file context
            FILE_CONTEXT_MAX = 30000
            file_context = ""
            attached_filenames: list[str] = []
            if files:
                from cscode.utils.file_parser import parse_file

                parts = []
                total_len = 0
                for name, content in files:
                    text = parse_file(name, content)
                    remaining = FILE_CONTEXT_MAX - total_len
                    if remaining <= 0:
                        parts.append(f"[FILE: {name}]\n[skipped: context limit reached]")
                        continue
                    if len(text) > remaining:
                        text = (
                            text[:remaining]
                            + f"\n[truncated: file too long, showing {remaining} of {len(text)} characters]"
                        )
                    total_len += len(text)
                    parts.append(f"[FILE: {name}]\n{text}")
                    attached_filenames.append(name)
                header = (
                    "The user attached the following file(s). "
                    "The full content is provided below in [FILE: ...] blocks. "
                    "Glob, Grep, and Ls are blocked for attached files. "
                    "Use Read tool if you need to read any file's content.\n\n"
                )
                file_context = "\n\n" + header + "\n\n".join(parts)

            # Build messages for LLM
            user_text = message.strip() if message else "请分析附件内容"

            # Append prompt.admitted event
            await _project_events(
                str(session_v2.session_id),
                [
                    {
                        "type": "prompt.admitted",
                        "data": {"prompt": message, "files": attached_filenames},
                    }
                ],
            )

            queue: asyncio.Queue[dict[str, object]] = asyncio.Queue()

            def _e(data: dict[str, object]) -> str:
                data["session_id"] = str(session_v2.session_id)
                return f"data: {json.dumps(data)}\n\n"

            async def on_event(event: LLMEvent) -> None:
                sse_event = _llm_event_to_dict(event)
                sse_event["session_id"] = str(session_v2.session_id)
                await queue.put(sse_event)
                evt_type = sse_event.get("type", "")
                if isinstance(evt_type, str) and evt_type in PERSIST_EVENT_TYPES:
                    evt_data = sse_event.get("data", {})
                    await _project_events(
                        str(session_v2.session_id),
                        [
                            {
                                "type": evt_type,
                                "data": dict(evt_data) if isinstance(evt_data, dict) else {},
                            }
                        ],
                    )
                # Notify TaskTracker for tool events with task_id
                if _tracker is not None and sse_event.get("type") in (
                    "tool.success",
                    "tool.failed",
                ):
                    await _tracker.handle_event(str(session_v2.session_id), sse_event)

            async def _emit_step_started() -> None:
                step_event: dict[str, object] = {
                    "type": "step.started",
                    "data": {},
                    "session_id": str(session_v2.session_id),
                }
                await queue.put(step_event)
                await _project_events(
                    str(session_v2.session_id), [{"type": "step.started", "data": {}}]
                )

            async def _emit_step_ended() -> None:
                step_event: dict[str, object] = {
                    "type": "step.ended",
                    "data": {},
                    "session_id": str(session_v2.session_id),
                }
                await queue.put(step_event)
                await _project_events(
                    str(session_v2.session_id), [{"type": "step.ended", "data": {}}]
                )

            before = time.time()

            # Create agent for this request
            if _db is not None:
                from cscode.core.config import Config, ConfigStore, load_config

                store = ConfigStore(_db)
                saved_config_raw = await store.get()
                if saved_config_raw is not None:
                    saved_config = Config.from_dict(saved_config_raw)
                else:
                    saved_config = load_config()
            else:
                from cscode.core.config import load_config

                saved_config = load_config()

            agent = create_agent_v2(saved_config, tool_registry=_tool_registry)
            agent._max_tool_rounds = 20

            async def agent_runner() -> str:
                logger.info(
                    "[DIAG] agent_runner: starting run_with_messages session=%s",
                    session_v2.session_id,
                )
                t0 = time.time()
                try:
                    # Build context messages
                    new_messages = await _build_context_messages(
                        session_v2, user_text, file_context
                    )

                    result = await agent.run_with_messages(
                        new_messages,
                        on_event=on_event,
                    )
                    return result
                finally:
                    logger.info(
                        "[DIAG] agent_runner: completed session=%s in %.1fs",
                        session_v2.session_id,
                        time.time() - t0,
                    )

            # Use Coordinator for per-session serialization
            async def run_with_coordinator() -> str:
                if _coordinator is not None:
                    return await _coordinator.run(
                        str(session_v2.session_id), _CallableProcessor(agent_runner)
                    )
                else:
                    return await agent_runner()

            # Cancel any existing agent task for this session
            old_task = _active_agent_tasks.get(str(session_v2.session_id))
            if old_task and not old_task.done():
                logger.info("Cancelling previous agent task for session %s", session_v2.session_id)
                old_task.cancel()
                try:
                    await asyncio.wait_for(old_task, timeout=5.0)
                except (asyncio.CancelledError, asyncio.TimeoutError):
                    pass

            _agent_event_queue: asyncio.Queue[dict[str, object]] = queue

            async def stepped_agent_runner() -> str:
                await _emit_step_started()
                try:
                    return await run_with_coordinator()
                finally:
                    await _emit_step_ended()

            agent_task = asyncio.create_task(stepped_agent_runner())
            _active_agent_tasks[str(session_v2.session_id)] = agent_task
            _session_queues[str(session_v2.session_id)] = queue

            last_event_time = time.time()
            last_status_time = time.time()
            response = ""

            while True:
                if await request.is_disconnected():
                    logger.info(
                        "Client disconnected for session %s, cancelling task", session_v2.session_id
                    )
                    if _question_registry is not None:
                        await _question_registry.cancel_session(str(session_v2.session_id))
                    agent_task.cancel()
                    break

                if agent_task.done():
                    # Drain remaining queue events
                    while not queue.empty():
                        try:
                            event = queue.get_nowait()
                            yield f"data: {json.dumps(event)}\n\n"
                        except asyncio.QueueEmpty:
                            break

                    # Auto-generate title if still using default (handles both
                    # direct SSE new sessions AND frontend pre-created sessions)
                    _current_title = session_v2.state.title
                    if not _current_title or _current_title in ("New Session", "New Chat"):
                        generated_title = None
                        try:
                            title_sys = NewMessage.system(
                                "You give very short session titles. Reply with ONLY 3-6 words."
                            )
                            if _db is not None:
                                from cscode.core.config import Config, ConfigStore, load_config

                                store = ConfigStore(_db)
                                saved_config_raw = await store.get()
                                if saved_config_raw is not None:
                                    saved_config = Config.from_dict(saved_config_raw)
                                else:
                                    saved_config = load_config()
                            else:
                                from cscode.core.config import load_config

                                saved_config = load_config()
                            title_agent = create_agent_v2(
                                saved_config, tool_registry=_tool_registry
                            )
                            title_r = await title_agent.llm_client.generate(
                                LLMRequest(
                                    model=title_agent.llm_client.route.model,
                                    messages=(title_sys, NewMessage.user(message or "")),
                                )
                            )
                            title_text = title_r.content.strip().strip("\"'.,!?")
                            if title_text:
                                generated_title = title_text
                        except Exception:
                            pass
                        if not generated_title:
                            generated_title = (
                                (message[:47] + "...")
                                if len(message) > 50
                                else (message or "New Chat")
                            )
                        await session_v2.update_metadata(title=generated_title)
                        yield _e({"type": "session:title", "data": {"title": generated_title}})

                    try:
                        response = agent_task.result()
                        for f in OUTPUTS_DIR.iterdir():
                            if f.is_file() and f.stat().st_mtime >= before:
                                yield _e({"type": "file_created", "data": {"filename": f.name}})
                        if not response:
                            logger.warning(
                                "[DIAG] Empty response from agent for session=%s",
                                session_v2.session_id,
                            )
                            yield _e(
                                {
                                    "type": "error",
                                    "data": {
                                        "content": "LLM returned empty response - API returned no content"
                                    },
                                }
                            )
                        else:
                            yield _e({"type": "complete", "data": {"content": response}})
                    except Exception as e:
                        logger.warning(
                            "[DIAG] agent task error session=%s error=%s", session_v2.session_id, e
                        )
                        detail = f"{type(e).__name__}: {e}\n{traceback.format_exc()}"
                        yield _e({"type": "error", "data": {"content": detail}})

                    asyncio.create_task(_auto_compact(str(session_v2.session_id), _event_store))
                    break

                try:
                    event = await asyncio.wait_for(queue.get(), timeout=0.5)
                    last_event_time = time.time()
                    yield f"data: {json.dumps(event)}\n\n"
                except asyncio.TimeoutError:
                    now = time.time()
                    if now - last_event_time > 10:
                        logger.info(
                            "[DIAG] keepalive session=%s last_event=%.1fs ago",
                            session_v2.session_id,
                            now - last_event_time,
                        )
                        yield ": keepalive\n\n"
                        last_event_time = now
                    if now - last_status_time > 60:
                        yield _e({"type": "status", "data": {"message": "Agent is working..."}})
                        last_status_time = now
                    continue

        except Exception as e:
            detail = f"{type(e).__name__}: {e}\n{traceback.format_exc()}"
            yield f"data: {json.dumps({'type': 'error', 'data': {'content': detail}, 'session_id': str(session_v2.session_id) if session_v2 else 'unknown'})}\n\n"

        finally:
            if session_v2 is not None:
                SessionLockManager.unlock(str(session_v2.session_id))
            if session_v2 is not None and agent_task is not None:
                if _active_agent_tasks.get(str(session_v2.session_id)) is agent_task:
                    del _active_agent_tasks[str(session_v2.session_id)]
                _session_queues.pop(str(session_v2.session_id), None)
                if not agent_task.done():
                    logger.info(
                        "[DIAG] Generator exiting, cancelling agent task for session %s",
                        session_v2.session_id,
                    )
                    agent_task.cancel()
                    try:
                        await asyncio.wait_for(agent_task, timeout=5.0)
                    except (asyncio.CancelledError, asyncio.TimeoutError):
                        pass

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@api_router.get("/events")
async def event_stream(session_id: str, after_seq: int = 0) -> StreamingResponse:
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    async def stream() -> AsyncIterator[str]:
        assert _event_store is not None
        import asyncio
        import json

        try:
            async for event in _event_store.subscribe(session_id, after_seq):
                yield f"event: {event.type}\ndata: {json.dumps(event.data)}\n\n"
        except asyncio.CancelledError:
            pass
        except Exception as e:
            yield f"event: error\ndata: {json.dumps({'message': str(e)})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@api_router.get("/sessions/{session_id}/events")
async def session_event_stream(session_id: str, after_seq: int = 0) -> StreamingResponse:
    """P1-3: Per-session SSE event subscription.

    Delegates to the existing /events endpoint logic for compatibility.
    """
    return await event_stream(session_id, after_seq)


# ── P2-2: WebSocket endpoint ──────────────────────────────────────

_ws_manager: WebSocketManager | None = None
_token_store: IntegrationTokenStore | None = None
_pty_manager: PTYSessionManager | None = None


@api_router.post("/integration/token")
async def create_integration_token(body: dict[str, str]) -> dict[str, object]:
    """Issue a temporary WebSocket auth token.

    Accepts an API key and returns a short-lived UUID token for WS auth.
    """
    global _token_store
    if _token_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    api_key = body.get("api_key", "")
    if not api_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    token = await _token_store.create(api_key)
    return {"token": token.token, "expires_at": int(token.expires_at)}


@api_router.post("/pty")
async def pty_endpoint(body: PTYInput) -> dict[str, object]:
    """POST /api/pty — PTY session management.

    Actions: create, exec, read, close, list.
    """
    global _pty_manager
    if _pty_manager is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    from cscode.tools2.pty import PTYTool

    tool = PTYTool(manager=_pty_manager)
    result = await tool.execute(body)

    if not result.success:
        raise HTTPException(status_code=400, detail=result.error)

    data = result.data
    if data is None:
        return {"success": True}
    if isinstance(data, list):
        return {"sessions": [s.model_dump() for s in data]}
    return data.model_dump()  # type: ignore[no-any-return]


async def _ws_chat_handler(client_id: str, message: dict[str, object]) -> None:
    """Bridge WebSocket chat messages to the chat system."""
    from cscode.server.state import state as app_state

    manager = app_state.ws_manager
    if manager is None:
        return

    await manager.send_to_client(client_id, {
        "type": "ack",
        "data": {"message": "message received, processing"},
    })

    session_id = message.get("session_id")
    if not isinstance(session_id, str):
        session_id = None
    msg_data = message.get("data", {})
    if isinstance(msg_data, dict):
        user_message = msg_data.get("message", "")
    else:
        user_message = str(msg_data)

    try:
        resp = await _handle_chat(user_message, session_id)
        await manager.send_to_client(client_id, {
            "type": "chat_response",
            "data": {
                "session_id": resp.session_id,
                "response": resp.response,
            },
        })
    except Exception:
        logger.exception("[WS] Chat handler error")
        await manager.send_to_client(client_id, {
            "type": "error",
            "data": {"message": "Chat processing failed"},
        })


@api_router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket) -> None:
    """P2-2: WebSocket endpoint for real-time bidirectional communication.

    Protocol:
      Client → Server: {"type": "ping"} | {"type": "subscribe", "session_id": "..."}
      Server → Client: {"type": "pong"} | {"type": "event", ...}

    See openspec/specs/cscode-integration-system.md for full protocol spec.
    """
    global _ws_manager
    manager = _ws_manager
    assert manager is not None, "_ws_manager should be initialised in lifespan"

    client = await manager.connect(websocket)
    try:
        await manager._handle_client_messages(client)
    finally:
        await manager.disconnect(client.client_id)


# ── P2-3: Workspace CRUD endpoints ─────────────────────────────────


class WorkspaceCreateRequest(BaseModel):
    name: str
    path: str
    config: dict[str, object] = {}


class WorkspaceUpdateRequest(BaseModel):
    name: str | None = None
    path: str | None = None
    config: dict[str, object] | None = None


@api_router.get("/workspaces")
async def list_workspaces(limit: int = 50) -> list[dict[str, object]]:
    """List all workspaces ordered by last_used_at descending."""
    global _workspace_store
    if _workspace_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    workspaces = await _workspace_store.list(limit=limit)
    return [_workspace_to_dict(ws) for ws in workspaces]


@api_router.get("/workspaces/recent")
async def recent_workspaces(limit: int = 10) -> list[dict[str, object]]:
    """List recently used workspaces."""
    global _workspace_store
    if _workspace_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    workspaces = await _workspace_store.recent(limit=limit)
    return [_workspace_to_dict(ws) for ws in workspaces]


@api_router.post("/workspaces")
async def create_workspace(req: WorkspaceCreateRequest) -> dict[str, object]:
    """Create a new workspace."""
    global _workspace_store
    if _workspace_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    try:
        ws = await _workspace_store.create(
            name=req.name,
            path=req.path,
            config=req.config,
        )
        return _workspace_to_dict(ws)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/workspaces/{workspace_id}")
async def get_workspace(workspace_id: str) -> dict[str, object]:
    """Get a workspace by id."""
    global _workspace_store
    if _workspace_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    ws = await _workspace_store.get(workspace_id)
    if ws is None:
        raise HTTPException(status_code=404, detail="Workspace not found")
    return _workspace_to_dict(ws)


@api_router.put("/workspaces/{workspace_id}")
async def update_workspace(workspace_id: str, req: WorkspaceUpdateRequest) -> dict[str, object]:
    """Update a workspace."""
    global _workspace_store
    if _workspace_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    try:
        ws = await _workspace_store.update(
            workspace_id=workspace_id,
            name=req.name,
            path=req.path,
            config=req.config,
        )
        if ws is None:
            raise HTTPException(status_code=404, detail="Workspace not found")
        return _workspace_to_dict(ws)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.delete("/workspaces/{workspace_id}", status_code=204)
async def delete_workspace(workspace_id: str) -> Response:
    """Delete a workspace."""
    global _workspace_store
    if _workspace_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    deleted = await _workspace_store.delete(workspace_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Workspace not found")
    return Response(status_code=204)


def _workspace_to_dict(ws: Any) -> dict[str, object]:
    """Convert a Workspace to a JSON-serializable dict."""
    return {
        "workspace_id": ws.workspace_id,
        "name": ws.name,
        "path": ws.path,
        "config": ws.config,
        "last_used_at": ws.last_used_at,
        "created_at": ws.created_at,
        "updated_at": ws.updated_at,
    }


@api_router.get("/workspaces/{workspace_id}/sessions")
async def list_workspace_sessions(workspace_id: str) -> list[dict[str, object]]:
    """P2-3: List all sessions associated with a workspace."""
    global _workspace_store, _event_store
    if _workspace_store is None or _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    ws = await _workspace_store.get(workspace_id)
    if ws is None:
        raise HTTPException(status_code=404, detail="Workspace not found")
    sessions = await _workspace_store.list_sessions(_event_store, workspace_id)
    return [
        {
            "session_id": str(s.session_id),
            "title": s.state.title,
            "model": s.state.model,
            "provider": s.state.provider,
            "status": s.state.status,
        }
        for s in sessions
    ]


@api_router.post("/sessions/{session_id}/move-workspace")
async def move_session_workspace(
    session_id: str,
    body: dict[str, str],
) -> dict[str, str]:
    """P2-4: Move a session to another workspace."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    to_ws = body.get("to_workspace_id", "")
    if not to_ws:
        raise HTTPException(status_code=400, detail="to_workspace_id is required")
    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    await session_v2.move_workspace(to_ws)
    return {"status": "ok"}


@api_router.get("/worktrees")
async def list_worktrees() -> list[dict[str, object]]:
    """P2-4: List all git worktrees."""
    from cscode.core.control_plane import WorktreeManager

    try:
        worktrees = WorktreeManager.list_worktrees()
        return [
            {
                "path": wt.path,
                "hash": wt.hash,
                "branch": wt.branch,
                "bare": wt.bare,
                "detached": wt.detached,
            }
            for wt in worktrees
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/worktrees")
async def create_worktree(body: dict[str, str]) -> dict[str, object]:
    """P2-4: Create a new git worktree."""
    from cscode.core.control_plane import WorktreeManager

    path = body.get("path", "")
    branch = body.get("branch", "")
    if not path:
        raise HTTPException(status_code=400, detail="path is required")
    success, message = WorktreeManager.add_worktree(path, branch or None)
    if not success:
        raise HTTPException(status_code=500, detail=message)
    return {"path": path, "message": message}


@api_router.delete("/worktrees")
async def remove_worktree(body: dict[str, str]) -> dict[str, str]:
    """P2-4: Remove (prune) a git worktree."""
    from cscode.core.control_plane import WorktreeManager

    path = body.get("path", "")
    if not path:
        raise HTTPException(status_code=400, detail="path is required")
    success, message = WorktreeManager.remove_worktree(path)
    if not success:
        raise HTTPException(status_code=500, detail=message)
    return {"status": "ok"}


async def _handle_chat(
    message: str, session_id: str | None, files: list[tuple[str, bytes]] | None = None
) -> ChatResponse:
    global _event_store, _tool_registry, _db
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_id = session_id or str(uuid.uuid4())

    # If session_id was explicitly provided, verify it exists
    if session_id:
        from cscode.schema.ids import SessionID as _SessionID

        existing_events = await _event_store.read(_SessionID(session_id))
        if not existing_events:
            raise HTTPException(status_code=404, detail="Session not found")

    try:
        # Load or create session
        config_data = None
        if _db is not None:
            from cscode.core.config import ConfigStore

            store = ConfigStore(_db)
            saved_config_dict = await store.get()
            if saved_config_dict:
                config_data = saved_config_dict

        session_v2, is_new = await _get_or_create_session(session_id, _event_store, config_data)

        # Build file context
        FILE_CONTEXT_MAX = 30000
        file_context = ""
        attached_filenames: list[str] = []
        if files:
            from cscode.utils.file_parser import parse_file

            parts = []
            total_len = 0
            for name, content in files:
                text = parse_file(name, content)
                remaining = FILE_CONTEXT_MAX - total_len
                if remaining <= 0:
                    parts.append(f"[FILE: {name}]\n[skipped: context limit reached]")
                    continue
                if len(text) > remaining:
                    text = (
                        text[:remaining]
                        + f"\n[truncated: file too long, showing {remaining} of {len(text)} characters]"
                    )
                total_len += len(text)
                parts.append(f"[FILE: {name}]\n{text}")
                attached_filenames.append(name)
            header = (
                "The user attached the following file(s). "
                "The full content is provided below in [FILE: ...] blocks. "
                "Glob, Grep, and Ls are blocked for attached files. "
                "Use Read tool if you need to read any file's content.\n\n"
            )
            file_context = "\n\n" + header + "\n\n".join(parts)

        # Build messages for LLM
        user_text = message.strip() if message else "请分析附件内容"
        messages = await _build_context_messages(session_v2, user_text, file_context)

        _compressor = ContextCompressor(threshold=50_000, keep_recent=10)
        if _compressor.needs_compression(messages):
            logger.info("Compressing %d messages before LLM call", len(messages))
            messages = _compressor.compress(messages)

        # Append prompt.admitted event
        await _project_events(
            str(session_v2.session_id),
            [{"type": "prompt.admitted", "data": {"prompt": message, "files": attached_filenames}}],
        )

        # Load permission rules for tool enforcement
        permission_rulesets = None
        if _db is not None:
            from cscode.app.factory import load_permission_rules
            permission_rulesets = await load_permission_rules(_db)

        # Create agent for this request
        if _db is not None:
            from cscode.core.config import Config, ConfigStore, load_config

            store = ConfigStore(_db)
            saved_config_raw = await store.get()
            if saved_config_raw is not None:
                saved_config = Config.from_dict(saved_config_raw)
            else:
                saved_config = load_config()
        else:
            from cscode.core.config import load_config

            saved_config = load_config()

        agent = create_agent_v2(saved_config, tool_registry=_tool_registry, permissions=permission_rulesets)

        # Persist LLM events to EventStore (matching streaming path behaviour)
        async def _on_event(event: LLMEvent) -> None:
            sse_event = _llm_event_to_dict(event)
            evt_type = sse_event.get("type", "")
            if isinstance(evt_type, str) and evt_type in PERSIST_EVENT_TYPES:
                evt_data = sse_event.get("data", {})
                await _project_events(
                    str(session_v2.session_id),
                    [
                        {
                            "type": evt_type,
                            "data": dict(evt_data) if isinstance(evt_data, dict) else {},
                        }
                    ],
                )

        # Run agent with full message history
        try:
            response = await agent.run_with_messages(messages, on_event=_on_event)
        except Exception as e:
            logger.error("agent_runner error: %s", e)
            raise

        # Auto-generate title if still using default
        _current_title = session_v2.state.title
        if not _current_title or _current_title in ("New Session", "New Chat"):
            generated_title = None
            try:
                title_sys = NewMessage.system(
                    "You give very short session titles. Reply with ONLY 3-6 words."
                )
                title_agent = create_agent_v2(saved_config, tool_registry=_tool_registry)
                title_r = await title_agent.llm_client.generate(
                    LLMRequest(
                        model=title_agent.llm_client.route.model,
                        messages=(title_sys, NewMessage.user(message or "")),
                    )
                )
                title_text = title_r.content.strip().strip("\"'.,!?")
                if title_text:
                    generated_title = title_text
            except Exception:
                pass
            if not generated_title:
                generated_title = (
                    (message[:47] + "...") if len(message) > 50 else (message or "New Chat")
                )
            await session_v2.update_metadata(title=generated_title)

        # Auto-compact if threshold exceeded (fire-and-forget)
        asyncio.create_task(_auto_compact(str(session_v2.session_id), _event_store))

        logger.debug("_handle_chat completed")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e

    return ChatResponse(response=response, session_id=session_id)


@api_router.get("/config")
async def get_config() -> dict[str, Any]:
    global _db
    if _db is not None:
        from cscode.core.config import ConfigStore

        store = ConfigStore(_db)
        saved_config = await store.get()
        if saved_config:
            saved_config.pop("api_key", None)
            return saved_config

    from cscode.core.config import load_config

    cfg = load_config()
    # Some PyInstaller builds may return dict instead of Config
    if isinstance(cfg, dict):
        cfg.pop("api_key", None)
        return cfg
    d = cfg.to_dict()
    d.pop("api_key", None)
    return d


@api_router.post("/config")
async def save_config(config: ConfigRequest) -> dict[str, str]:
    global _db
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    from cscode.core.config import ConfigStore

    store = ConfigStore(_db)
    await store.save(config.model_dump())

    return {"status": "ok"}


@api_router.put("/config")
async def update_config(config: ConfigRequest) -> dict[str, str]:
    """Alias for POST /config — accepts PUT for cross-tool compatibility."""
    return await save_config(config)


@api_router.post("/sessions/{session_id}/export")
async def export_session(session_id: str) -> Response:
    global _event_store, _projector
    if _event_store is None or _projector is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    # Load session for validation (events are used for state projection)
    await SessionV2.load(_event_store, SessionID(session_id))  # noqa: F841
    events = await _event_store.read(session_id)
    state = SessionProjector.project(events)

    import json

    data = {
        "session_id": str(state.session_id),
        "title": state.title,
        "provider": state.provider,
        "model": state.model,
        "created_at": state.created_at,
        "updated_at": state.updated_at,
        "messages": [
            {
                "role": msg.role,
                "content": msg.content,
            }
            for msg in state.messages
        ],
    }
    from urllib.parse import quote

    safe_filename = state.title.replace(" ", "_")
    encoded_filename = quote(safe_filename, safe="", encoding="utf-8")
    return Response(
        content=json.dumps(data, indent=2, ensure_ascii=False),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}.json"},
    )


@api_router.post("/sessions/import")
async def import_session(request: Request) -> dict[str, str]:
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    body = await request.json()

    # Create new session with imported data
    session_v2 = await SessionV2.create(
        _event_store,
        model=body.get("model", "gpt-4o"),
        provider=body.get("provider", "openai"),
        title=body.get("title", "Imported Session"),
    )

    # Replay messages as events
    for msg in body.get("messages", []):
        if msg.get("role") == "user":
            await _project_events(
                str(session_v2.session_id),
                [{"type": "prompt.admitted", "data": {"prompt": msg.get("content", "")}}],
            )
        elif msg.get("role") == "assistant":
            await _project_events(
                str(session_v2.session_id),
                [{"type": "text.ended", "data": {"content": msg.get("content", "")}}],
            )

    return {"id": str(session_v2.session_id), "title": session_v2.state.title}


def _format_dt(dt: object) -> str | None:
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat()
    return str(dt)


@api_router.get("/share")
async def list_shares(session_id: str | None = None) -> list[dict[str, object]]:
    global _share_store
    if _share_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    if session_id:
        shares = await _share_store.list_by_session(session_id)
    else:
        shares = await _share_store.list()
    return [
        {
            "id": s.id,
            "session_id": s.session_id,
            "title": s.title,
            "created_at": _format_dt(s.created_at),
            "expires_at": _format_dt(s.expires_at) if s.expires_at else None,
            "is_active": s.is_active,
        }
        for s in shares
    ]


@api_router.post("/share", status_code=201)
async def create_share(request: Request) -> dict[str, str]:
    global _share_store
    if _share_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    body = await request.json()
    share = await _share_store.create(
        session_id=body.get("session_id", ""),
        title=body.get("title", ""),
    )
    return {"id": share.id}


@api_router.get("/share/{share_id}")
async def get_share(share_id: str) -> dict[str, object]:
    global _share_store
    if _share_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    share = await _share_store.get(share_id)
    if share is None:
        raise HTTPException(status_code=404, detail="Share not found")
    return {
        "id": share.id,
        "session_id": share.session_id,
        "title": share.title,
        "created_at": _format_dt(share.created_at),
        "expires_at": _format_dt(share.expires_at) if share.expires_at else None,
        "is_active": share.is_active,
    }


@api_router.delete("/share/{share_id}", status_code=204)
async def delete_share(share_id: str) -> None:
    global _share_store
    if _share_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    result = await _share_store.delete(share_id)
    if not result:
        raise HTTPException(status_code=404, detail="Share not found")


_lsp_manager: LSPManager | None = None


@api_router.get("/lsp/diagnostics")
async def get_lsp_diagnostics(file_path: str) -> dict[str, object]:
    """Get LSP diagnostics for a file. Used by frontend to show errors/warnings."""
    global _lsp_manager
    if _lsp_manager is None:
        _lsp_manager = LSPManager()
    try:
        from cscode.tools2.lsp import LSPInput, LSPTool

        tool = LSPTool(_lsp_manager)
        result = await tool.execute(
            LSPInput(
                command="diagnostics",
                file_path=file_path,
            )
        )
        if not result.success:
            return {"results": [], "error": result.error}
        diagnostics = result.data.results if result.data else []
        return {"results": diagnostics}
    except Exception as e:
        return {"results": [], "error": str(e)}


# ── OAuth callback (for MCP browser-based auth flow) ─────────────────


_oAuth_codes: dict[str, dict[str, str]] = {}


@api_router.get("/auth/callback")
async def oauth_callback(
    code: str | None = None,
    state: str | None = None,
    iss: str | None = None,
    error: str | None = None,
) -> dict[str, str]:
    """OAuth authorization callback endpoint.

    The authorization server redirects the browser here after the user
    grants/denies access. Stores the authorization code for the MCP
    OAuth client to retrieve.
    """
    if error:
        return {"status": "error", "error": error}

    if code is None or state is None:
        raise HTTPException(status_code=400, detail="Missing code or state parameter")

    _oAuth_codes[state] = {"code": code, "iss": iss or ""}
    return {"status": "success", "message": "Authorization received, you can close this window"}


@api_router.get("/auth/token")
async def get_oauth_token(state: str) -> dict[str, str]:
    """Retrieve stored OAuth authorization code by state."""
    data = _oAuth_codes.pop(state, None)
    if data is None:
        raise HTTPException(status_code=404, detail="State not found or already consumed")
    return data


@api_router.get("/sessions/{session_id}/messages")
async def get_session_messages(session_id: str) -> list[dict[str, object]]:
    """P0-1: Return messages for a session (used by sidebar session switching)."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    messages = SessionProjector.build_context(session_v2.state)
    return [
        {
            "role": msg.role,
            "content": msg.content,
            "id": str(msg.id) if msg.id is not None else _make_msg_id(msg.role, msg.content, i),
        }
        for i, msg in enumerate(messages)
    ]


def _make_msg_id(role: str, content: str, index: int) -> str:
    """Generate a stable synthetic message ID from role + content hash."""
    import hashlib

    raw = f"{role}:{content}:{index}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


@api_router.get("/sessions/{session_id}/context")
async def get_session_context(session_id: str) -> list[dict[str, object]]:
    """P1-2: Return LLM context messages for a session (with system prompts)."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    messages = SessionProjector.build_context(session_v2.state)
    return [{"role": msg.role, "content": msg.content} for msg in messages]


@api_router.get("/sessions/{session_id}/summary")
async def get_session_summary(session_id: str) -> dict[str, object]:
    """P1-8: Return a statistical summary of a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    from cscode.core.session_summary import SessionSummary

    return SessionSummary(session_v2).generate()


@api_router.post("/sessions/{session_id}/model")
async def switch_model(session_id: str, body: dict[str, object]) -> dict[str, str]:
    """P1-5: Switch the model/provider for a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    model = body.get("model", session_v2.state.model)
    provider = body.get("provider", session_v2.state.provider)
    await session_v2.update_metadata(model=str(model))
    logger.info("Session %s model switched to %s (provider=%s)", session_id, model, provider)
    return {"status": "ok"}


@api_router.post("/sessions/{session_id}/agent")
async def switch_agent(session_id: str, body: dict[str, object]) -> dict[str, str]:
    """P1-5: Switch the agent for a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    agent = body.get("agent", "auto")
    await session_v2.update_metadata(agent=str(agent))
    logger.info("Session %s agent switched to %s", session_id, agent)
    return {"status": "ok"}


@api_router.post("/sessions/{session_id}/workspace")
async def associate_session_workspace(
    session_id: str,
    body: dict[str, str],
) -> dict[str, str]:
    """P2-3: Associate a session with a workspace."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    ws_id = body.get("workspace_id", "")
    if not ws_id:
        raise HTTPException(status_code=400, detail="workspace_id is required")
    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    await session_v2.associate_workspace(ws_id)
    return {"status": "ok"}


@api_router.get("/sessions/{session_id}/info")
async def get_session_info(session_id: str) -> dict[str, object]:
    """P2-7: Return full session metadata."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    st = session_v2.state
    return {
        "session_id": str(st.session_id),
        "title": st.title,
        "model": st.model,
        "provider": st.provider,
        "agent": st.agent,
        "status": st.status,
        "workspace_id": st.workspace_id,
        "message_count": len(st.messages),
        "event_count": st.seq,
        "tool_rounds": st.tool_rounds,
        "created_at": st.created_at,
        "updated_at": st.updated_at,
        "seq": st.seq,
    }


@api_router.get("/sessions/{session_id}/instruction")
async def get_session_instruction(session_id: str) -> dict[str, str]:
    """P2-6: Get the per-session custom instruction."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    return {"instruction": session_v2.state.instruction}


@api_router.put("/sessions/{session_id}/instruction")
async def set_session_instruction(session_id: str, body: dict[str, object]) -> dict[str, str]:
    """P2-6: Set or update the per-session custom instruction."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    instruction = str(body.get("instruction", ""))
    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    await session_v2.set_instruction(instruction)
    return {"instruction": session_v2.state.instruction}


@api_router.delete("/sessions/{session_id}/instruction")
async def delete_session_instruction(session_id: str) -> dict[str, bool]:
    """P2-6: Remove the per-session custom instruction."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    await session_v2.delete_instruction()
    return {"deleted": True}


@api_router.get("/sessions/{session_id}/run-state")
async def get_session_run_state(session_id: str) -> dict[str, str]:
    """P2-9: Get the current run state of a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    return {
        "status": session_v2.state.run_status,
        "error": session_v2.state.run_error,
    }


@api_router.put("/sessions/{session_id}/run-state")
async def set_session_run_state(session_id: str, body: dict[str, str]) -> dict[str, str]:
    """P2-9: Set the run state of a session.

    Valid status values: running, stopped, errored, completed.
    """
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    status = body.get("status", "")
    if status not in ("running", "stopped", "errored", "completed"):
        raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    method_map = {
        "running": session_v2.mark_run_start,
        "stopped": session_v2.mark_run_stop,
        "errored": lambda: session_v2.mark_run_error(error=body.get("error", "")),
        "completed": session_v2.mark_run_complete,
    }
    fn = method_map[status]
    await fn()

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    return {
        "status": session_v2.state.run_status,
        "error": session_v2.state.run_error,
    }


@api_router.get("/config/reference")
async def get_config_reference() -> list[dict[str, str]]:
    """P2-10: Return schema of all known config keys with types, defaults, descriptions."""
    from cscode.core.config import CONFIG_KEY_META

    return [
        {
            "key": k,
            "type": v.get("type", "string"),
            "default": v.get("default", ""),
            "description": v.get("description", ""),
        }
        for k, v in sorted(CONFIG_KEY_META.items())
    ]


@api_router.get("/tools/application")
async def list_application_tools() -> dict[str, list[str]]:
    """List all application-level tools (safe, read-only tools)."""
    from cscode.core.application_tools import get_application_tools

    return {"tools": get_application_tools()}


@api_router.get("/tools")
async def list_all_tools() -> dict[str, list[str]]:
    """Alias for /tools/application — list all available tools."""
    from cscode.core.application_tools import get_application_tools

    return {"tools": get_application_tools()}


@api_router.get("/version")
async def get_version() -> dict[str, str]:
    """Return the application version."""
    from cscode import __version__

    return {"version": __version__, "app": "CScode"}


@api_router.get("/sessions/{session_id}/overflow")
async def get_session_overflow(session_id: str) -> dict[str, bool | int]:
    """P2-12: Check if a session is overflowing (too many messages)."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    info = session_v2.check_overflow()
    return {
        "overflowing": info["overflowing"],
        "near_overflow": info["near_overflow"],
        "message_count": info["message_count"],
        "threshold": info["threshold"],
    }


@api_router.post("/sessions/{session_id}/retry")
async def retry_session(session_id: str) -> dict[str, bool | str | int]:
    """P2-13: Retry the last prompt in a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    last = session_v2.get_last_prompt()
    if last is None:
        raise HTTPException(status_code=400, detail="No prompt to retry")

    events = await session_v2.retry()
    return {
        "retried": True,
        "last_prompt": last,
        "event_count": len(events),
    }


@api_router.get("/sessions/{session_id}/reminders")
async def list_reminders(session_id: str) -> dict[str, list[dict[str, object]]]:
    """P2-14: List all reminders for a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    return {"reminders": session_v2.state.reminders}


@api_router.post("/sessions/{session_id}/reminders")
async def add_reminder(session_id: str, body: dict[str, str]) -> dict[str, object]:
    """P2-14: Add a reminder to a session."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")

    text = body.get("text", "")
    if not text:
        raise HTTPException(status_code=400, detail="text is required")

    session_v2 = await SessionV2.load(_event_store, SessionID(session_id))
    if session_v2.state.seq == 0:
        raise HTTPException(status_code=404, detail="Session not found")

    return await session_v2.add_reminder(text)


@api_router.get("/directories/external")
async def list_external_directories() -> dict[str, list[dict[str, object]]]:
    """P2-16: List all approved external directories."""
    global _external_dir_store
    if _external_dir_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    return {
        "directories": [
            {"id": d.id, "path": d.path, "created_at": d.created_at}
            for d in _external_dir_store.list()
        ]
    }


@api_router.post("/directories/external")
async def add_external_directory(body: dict[str, str]) -> dict[str, object]:
    """P2-16: Register a new approved external directory."""
    global _external_dir_store
    if _external_dir_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    path = body.get("path", "")
    if not path:
        raise HTTPException(status_code=400, detail="path is required")
    try:
        entry = _external_dir_store.add(path)
        return {"id": entry.id, "path": entry.path, "created_at": entry.created_at}
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))


@api_router.delete("/directories/external/{dir_id}")
async def remove_external_directory(dir_id: str) -> dict[str, bool]:
    """P2-16: Remove an approved external directory."""
    global _external_dir_store
    if _external_dir_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    ok = _external_dir_store.remove(dir_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Directory not found")
    return {"ok": True}


@api_router.get("/directories/external/check")
async def check_external_directory(path: str = "") -> dict[str, bool]:
    """P2-16: Check if a path is within an approved external directory."""
    global _external_dir_store
    if _external_dir_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    return {"approved": _external_dir_store.is_approved(path)}


@api_router.get("/sessions/{session_id}/questions")
async def list_questions(session_id: str) -> list[dict[str, object]]:
    """P0-2: List pending questions for a session (used by frontend polling)."""
    global _question_registry
    if _question_registry is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    return await _question_registry.list_pending(session_id)


@api_router.post("/sessions/{session_id}/questions/{request_id}/reply")
async def reply_question(
    session_id: str, request_id: str, body: dict[str, object]
) -> dict[str, str]:
    """P0-2: Reply to a pending question."""
    global _question_registry
    if _question_registry is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    answers = body.get("answers", [])
    if isinstance(answers, list):
        str_answers = [str(a) for a in answers]
    else:
        str_answers = [str(answers)]
    ok = await _question_registry.resolve(request_id, str_answers)
    if not ok:
        raise HTTPException(status_code=404, detail="Question not found or already answered")

    # P2-3: If always_allow is true, auto-save a global allow-all permission rule (SQLite)
    if body.get("always_allow") and state.db is not None:
        logger.info("always_allow triggered for session=%s request=%s", session_id, request_id)
        try:
            from cscode.core.permission_v2 import Rule as PermissionRule, RuleEffect, SavedRules as PermissionSavedRules

            saved = PermissionSavedRules(state.db)
            await saved.save(PermissionRule(action="*", resource="*", effect=RuleEffect.ALLOW))
        except Exception:
            logger.exception("Failed to save always_allow rule")

    return {"status": "ok"}


@api_router.post("/sessions/{session_id}/questions/{request_id}/reject")
async def reject_question(session_id: str, request_id: str) -> dict[str, str]:
    """P0-2: Reject a pending question."""
    global _question_registry
    if _question_registry is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    ok = await _question_registry.reject(request_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Question not found or already answered")
    return {"status": "ok"}


@api_router.get("/files/search")
async def search_files(q: str = "") -> list[str]:
    try:
        from cscode.tools.glob import GlobTool

        tool = GlobTool()
        result = await tool.execute({"pattern": q if q else "**/*"})
        if result.success and result.data:
            if "No files matching" in result.data:
                return []
            files = result.data.split("\n")
            return [f for f in files if f]
    except Exception:
        pass
    return []


# ---------------------------------------------------------------------------
# P2-4: Session compact API
# ---------------------------------------------------------------------------


@api_router.post("/sessions/{session_id}/compact")
async def compact_session(session_id: str) -> dict[str, object]:
    """Compress a session by replacing old events with a summary."""
    global _compactor
    if _compactor is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    try:
        baseline_seq = await _compactor.compact(session_id)
        return {"status": "ok", "baseline_seq": baseline_seq}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# P2-5: File system API extensions
# ---------------------------------------------------------------------------


class FileReadRequest(BaseModel):
    path: str


@api_router.post("/files/read")
async def read_file(req: FileReadRequest) -> dict[str, object]:
    """Read a file's contents from the workspace."""
    import os

    try:
        path = os.path.abspath(os.path.expanduser(req.path))
        if not os.path.isfile(path):
            raise HTTPException(status_code=404, detail="File not found")
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        return {"path": path, "content": content, "size": len(content)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/files/list")
async def list_directory(path: str = ".") -> dict[str, object]:
    """List contents of a directory."""
    import os
    import stat as stat_module

    try:
        dir_path = os.path.abspath(os.path.expanduser(path))
        if not os.path.isdir(dir_path):
            raise HTTPException(status_code=404, detail="Directory not found")
        entries: list[dict[str, object]] = []
        for name in sorted(os.listdir(dir_path)):
            full = os.path.join(dir_path, name)
            try:
                st = os.stat(full)
                mode = st.st_mode
                entries.append(
                    {
                        "name": name,
                        "type": "dir"
                        if stat_module.S_ISDIR(mode)
                        else "file"
                        if stat_module.S_ISREG(mode)
                        else "other",
                        "size": st.st_size,
                        "mtime": st.st_mtime,
                    }
                )
            except OSError:
                entries.append({"name": name, "type": "unknown", "size": 0, "mtime": 0})
        return {"path": dir_path, "entries": entries, "count": len(entries)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# P0-3: Session Input Inbox — event-sourced input queue per session
# ---------------------------------------------------------------------------


@api_router.get("/sessions/{session_id}/inbox")
async def get_inbox(session_id: str) -> dict[str, object]:
    """Get the current inbox state (pending inputs, processing ID)."""
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _event_store is not None
    from cscode.core.session_input import InputInbox

    inbox = InputInbox(_event_store, session_id)
    await inbox.reload()
    return {
        "pending": [
            {"id": inp.id, "content": inp.content[:200], "created_at": inp.created_at}
            for inp in inbox.state.pending
        ],
        "processing_id": inbox.state.processing_id,
    }


@api_router.post("/sessions/{session_id}/inbox", status_code=201)
async def enqueue_input(session_id: str, request: Request) -> dict[str, object]:
    """Enqueue a new input to the session's inbox."""
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _event_store is not None
    from cscode.core.session_input import InputInbox

    body = await request.json()
    inbox = InputInbox(_event_store, session_id)
    try:
        inp = await inbox.enqueue(
            content=body.get("content", ""),
            files=body.get("files"),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    return {"id": inp.id, "content": inp.content[:200]}


@api_router.delete("/sessions/{session_id}/inbox/{input_id}")
async def cancel_input(session_id: str, input_id: str) -> dict[str, bool]:
    """Cancel a pending input in the session's inbox."""
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _event_store is not None
    from cscode.core.session_input import InputInbox

    inbox = InputInbox(_event_store, session_id)
    await inbox.reload()
    cancelled = await inbox.cancel(input_id)
    return {"cancelled": cancelled}


@api_router.delete("/sessions/{session_id}/inbox", status_code=204)
async def clear_inbox(session_id: str) -> None:
    """Clear all pending inputs in the session's inbox."""
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _event_store is not None
    from cscode.core.session_input import InputInbox

    inbox = InputInbox(_event_store, session_id)
    await inbox.reload()
    await inbox.clear()


# ---------------------------------------------------------------------------
# Task Verification Report
# ---------------------------------------------------------------------------


@api_router.get("/sessions/{session_id}/verification-report")
async def get_verification_report(session_id: str) -> dict[str, object]:
    """Return a database-backed verification report for a session.

    The report is generated from the task_verifications projection table,
    not from LLM text. This prevents the LLM from fabricating results.
    """
    global _tracker
    if _tracker is None:
        return {
            "summary": {"executed": 0, "failed": 0, "unverified": 0, "skipped": 0},
            "details": [],
        }
    report = await _tracker.get_execution_report(session_id)

    # Compute SKIPPED: expected tasks not in the projection table
    db = getattr(_tracker, "db", None)
    if db is not None:
        all_expected = await db.fetchall(
            "SELECT task_id FROM expected_tasks WHERE session_id = ?",
            (session_id,),
        )
        expected_ids = {r["task_id"] for r in all_expected}
        recorded_ids = {d["task_id"] for d in report["details"]}
        skipped = expected_ids - recorded_ids
        report["summary"]["skipped"] = len(skipped)
        report["details"].extend(
            [
                {
                    "task_id": tid,
                    "status": "SKIPPED",
                    "evidence": {},
                    "result_summary": "",
                    "timestamp": None,
                }
                for tid in skipped
            ]
        )
    return report


# ---------------------------------------------------------------------------
# P0-8: Provider Status — check LLM provider availability
# ---------------------------------------------------------------------------


@api_router.get("/providers/status")
async def get_provider_status(provider: str = "") -> dict[str, object]:
    """Check the status of LLM providers.

    If provider is specified, checks that provider only.
    Otherwise checks all known providers.
    """
    from cscode.providers.status import ProviderStatusChecker

    checker = ProviderStatusChecker()

    if provider:
        result = checker.check(provider)
        return {"provider": provider, "status": result.status.value, "message": result.message}

    from cscode.providers.status import _DEFAULT_BASE_URLS

    results: dict[str, dict[str, str]] = {}
    for p in _DEFAULT_BASE_URLS:
        result = checker.check(p)
        results[p] = {"status": result.status.value, "message": result.message}
    return {"providers": results}


# ---------------------------------------------------------------------------
# P1-1: Credential CRUD — secure credential storage
# ---------------------------------------------------------------------------


@api_router.get("/credentials")
async def list_credentials(
    provider: str | None = None,
    cred_type: str | None = None,
) -> list[dict[str, object]]:
    """List stored credentials with optional filters."""
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _db is not None
    from cscode.core.credential import CredentialStore

    store = CredentialStore(_db)
    creds = await store.list(provider=provider, cred_type=cred_type)
    return [
        {
            "id": c.id,
            "name": c.name,
            "type": c.type,
            "display_value": c.display_value,
            "provider": c.provider,
            "created_at": c.created_at,
            "updated_at": c.updated_at,
            "expires_at": c.expires_at,
            "is_expired": c.is_expired,
        }
        for c in creds
    ]


@api_router.post("/credentials", status_code=201)
async def create_credential(request: Request) -> dict[str, str]:
    """Create a new credential."""
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _db is not None
    from cscode.core.credential import CredentialStore

    body = await request.json()
    store = CredentialStore(_db)
    cred = await store.create(
        name=body.get("name", ""),
        type=body.get("type", "api_key"),
        value=body.get("value", ""),
        provider=body.get("provider", "custom"),
        expires_at=body.get("expires_at"),
    )
    return {"id": cred.id}


@api_router.get("/credentials/{cred_id}")
async def get_credential(cred_id: str) -> dict[str, object]:
    """Get a credential by ID (with masked display value)."""
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _db is not None
    from cscode.core.credential import CredentialStore

    store = CredentialStore(_db)
    cred = await store.get(cred_id)
    if cred is None:
        raise HTTPException(status_code=404, detail="Credential not found")
    return {
        "id": cred.id,
        "name": cred.name,
        "type": cred.type,
        "display_value": cred.display_value,
        "provider": cred.provider,
        "created_at": cred.created_at,
        "updated_at": cred.updated_at,
        "expires_at": cred.expires_at,
        "is_expired": cred.is_expired,
    }


@api_router.put("/credentials/{cred_id}")
async def update_credential(cred_id: str, request: Request) -> dict[str, object]:
    """Update a credential's mutable fields."""
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _db is not None
    from cscode.core.credential import CredentialStore

    body = await request.json()
    store = CredentialStore(_db)
    cred = await store.update(
        cred_id=cred_id,
        name=body.get("name"),
        value=body.get("value"),
        expires_at=body.get("expires_at"),
    )
    if cred is None:
        raise HTTPException(status_code=404, detail="Credential not found")
    return {"id": cred.id, "updated_at": cred.updated_at}


@api_router.delete("/credentials/{cred_id}", status_code=204)
async def delete_credential(cred_id: str) -> None:
    """Delete a credential by ID."""
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _db is not None
    from cscode.core.credential import CredentialStore

    store = CredentialStore(_db)
    deleted = await store.delete(cred_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Credential not found")


@api_router.post("/credentials/{cred_id}/rotate")
async def rotate_credential(cred_id: str, request: Request) -> dict[str, object]:
    """Rotate a credential value, preserving the previous value."""
    if _db is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    assert _db is not None
    from cscode.core.credential import CredentialStore

    body = await request.json()
    store = CredentialStore(_db)
    try:
        cred = await store.rotate(cred_id, body.get("new_value", ""))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if cred is None:
        raise HTTPException(status_code=404, detail="Credential not found")
    return {"id": cred.id, "rotated_at": cred.rotated_at}


# ---------------------------------------------------------------------------
# P1-4: Catalog — model/provider/agent registry
# ---------------------------------------------------------------------------


_catalog: object | None = None


def _get_catalog() -> object:
    """Lazy-initialized singleton catalog with known providers."""
    global _catalog
    if _catalog is not None:
        return _catalog
    from cscode.core.catalog import (
        AgentEntry,
        Catalog,
        ModelEntry,
        ProviderEntry,
    )

    cat = Catalog()
    # Pre-populate known providers
    providers_data = [
        ("openai", "OpenAI", "openai"),
        ("anthropic", "Anthropic", "anthropic"),
        ("gemini", "Google Gemini", "gemini"),
        ("azure", "Azure OpenAI", "openai"),
        ("ollama", "Ollama", "openai"),
        ("openrouter", "OpenRouter", "openai"),
        ("bedrock", "AWS Bedrock", "anthropic"),
        ("cohere", "Cohere", "openai"),
        ("grok", "Grok (xAI)", "openai"),
        ("mistral", "Mistral AI", "openai"),
        ("nvidia", "NVIDIA NIM", "openai"),
        ("perplexity", "Perplexity", "openai"),
        ("vertex", "Google Vertex AI", "vertex"),
        ("xai", "xAI", "openai"),
    ]
    models_data = [
        ("gpt-4o", "GPT-4o", "openai", ["chat", "vision"], 128000),
        ("gpt-4o-mini", "GPT-4o Mini", "openai", ["chat", "vision"], 128000),
        ("claude-sonnet-4-20250514", "Claude Sonnet 4", "anthropic", ["chat"], 200000),
        ("claude-haiku-3-5-20241022", "Claude Haiku 3.5", "anthropic", ["chat"], 200000),
        ("gemini-2.5-pro-exp-03-25", "Gemini 2.5 Pro", "gemini", ["chat"], 1048576),
        ("command-a-03-2025", "Command A", "cohere", ["chat"], 256000),
        ("deepseek-chat", "DeepSeek V3", "openrouter", ["chat"], 65536),
    ]
    for pid, pname, api_type in providers_data:
        cat.register_provider(ProviderEntry(id=pid, name=pname, api_type=api_type))
    for mid, mname, mprovider, caps, ctx in models_data:
        cat.register_model(
            ModelEntry(
                id=mid, name=mname, provider=mprovider, capabilities=caps, context_length=ctx
            )
        )
    cat.register_agent(
        AgentEntry(
            id="default",
            name="Default Agent",
            description="General-purpose coding assistant",
            tools=["read", "grep", "edit", "bash"],
        )
    )
    # Sync agents from AgentRegistry so they appear in the catalog endpoint
    from cscode.core.agent.factory import get_registry
    reg = get_registry()
    for ad in reg.list():
        if not cat.get_agent(ad.name):
            cat.register_agent(AgentEntry(
                id=ad.name,
                name=ad.name,
                description=ad.description,
                tools=sorted(ad.capabilities),
            ))
    _catalog = cat
    return cat


@api_router.get("/catalog/models")
async def list_catalog_models(
    provider: str | None = None, search: str = ""
) -> list[dict[str, object]]:
    """List models in the catalog, optionally filtered by provider or search."""
    from cscode.core.catalog import Catalog

    cat: Catalog = _get_catalog()  # type: ignore[assignment]
    if search:
        results = cat.search_models(search)
    else:
        results = cat.list_models(provider=provider)
    return [
        {
            "id": m.id,
            "name": m.name,
            "provider": m.provider,
            "capabilities": m.capabilities,
            "context_length": m.context_length,
        }
        for m in results
    ]


@api_router.get("/catalog/providers")
async def list_catalog_providers() -> list[dict[str, object]]:
    """List all providers in the catalog."""
    from cscode.core.catalog import Catalog

    cat: Catalog = _get_catalog()  # type: ignore[assignment]
    return [
        {"id": p.id, "name": p.name, "api_type": p.api_type, "models": p.models}
        for p in cat.list_providers()
    ]


@api_router.get("/catalog/agents")
async def list_catalog_agents() -> list[dict[str, object]]:
    """List all agents in the catalog."""
    from cscode.core.catalog import Catalog

    cat: Catalog = _get_catalog()  # type: ignore[assignment]
    return [
        {"id": a.id, "name": a.name, "description": a.description, "tools": a.tools}
        for a in cat.list_agents()
    ]


# ---------------------------------------------------------------------------
# P1-11: Background Jobs — async job queue with status tracking
# ---------------------------------------------------------------------------

_job_queue: object | None = None


def _get_job_queue() -> object:
    global _job_queue
    if _job_queue is not None:
        return _job_queue
    from cscode.core.background_job import BackgroundJobQueue

    q = BackgroundJobQueue()
    _job_queue = q
    return q


@api_router.post("/jobs", status_code=201)
async def enqueue_job(request: Request) -> dict[str, str]:
    """Enqueue a new background job."""
    body = await request.json()
    from cscode.core.background_job import BackgroundJobQueue

    q: BackgroundJobQueue = _get_job_queue()  # type: ignore[assignment]
    job = await q.enqueue(
        job_type=body.get("job_type", ""),
        params=body.get("params"),
    )
    return {"id": job.id}


@api_router.get("/jobs")
async def list_jobs(
    status: str | None = None,
    job_type: str | None = None,
    limit: int = 50,
) -> list[dict[str, object]]:
    """List background jobs with optional filters."""
    from cscode.core.background_job import BackgroundJobQueue, JobStatus

    q: BackgroundJobQueue = _get_job_queue()  # type: ignore[assignment]
    status_enum = JobStatus(status) if status else None
    jobs = await q.list_jobs(status=status_enum, job_type=job_type, limit=limit)
    return [j.to_dict() for j in jobs]


@api_router.get("/jobs/{job_id}")
async def get_job(job_id: str) -> dict[str, object]:
    """Get the status and details of a background job."""
    from cscode.core.background_job import BackgroundJobQueue

    q: BackgroundJobQueue = _get_job_queue()  # type: ignore[assignment]
    job = await q.get_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return job.to_dict()


@api_router.post("/jobs/{job_id}/cancel")
async def cancel_job(job_id: str) -> dict[str, bool]:
    """Cancel a pending or running job."""
    from cscode.core.background_job import BackgroundJobQueue

    q: BackgroundJobQueue = _get_job_queue()  # type: ignore[assignment]
    cancelled = await q.cancel_job(job_id)
    return {"cancelled": cancelled}


# ---------------------------------------------------------------------------
# P0-4 / P1-12: File Attachments & Locale — convenience endpoints
# ---------------------------------------------------------------------------


@api_router.post("/files/attach")
async def create_attachment(request: Request) -> dict[str, object]:
    """Create an Attachment from a file path. Reads the file and returns metadata."""
    from cscode.core.attachment import Attachment

    body = await request.json()
    file_path = body.get("path", "")
    if not file_path:
        raise HTTPException(status_code=400, detail="path is required")
    try:
        attachment = Attachment.from_path(file_path)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"File not found: {file_path}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {
        "name": attachment.name,
        "path": attachment.path,
        "size": attachment.size,
        "mime_type": attachment.mime_type,
        "is_image": attachment.is_image,
        "is_text": attachment.is_text,
    }


@api_router.get("/locale")
async def get_locale() -> dict[str, str]:
    """Get the current system locale."""
    from cscode.core.i18n import I18n

    locale = I18n.detect_locale()
    return {"locale": locale}


@api_router.post("/locale")
async def set_locale(request: Request) -> dict[str, str]:
    """Set the locale (en or zh)."""
    from cscode.core.i18n import get_i18n

    body = await request.json()
    locale = body.get("locale", "en")
    get_i18n().set_locale(locale)
    return {"locale": get_i18n().locale}


# ─── P2-5: Sync — multi-instance event sync ──────────────────────────


@api_router.get("/sync/events")
async def get_sync_events(after_id: int = 0) -> list[dict[str, object]]:
    """P2-5: Return events with id > after_id for incremental sync."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    events = await _event_store.scan_events_global(after_id=after_id)
    return [
        {
            "id": e.id,
            "aggregate_id": e.aggregate_id,
            "seq": e.seq,
            "type": e.type,
            "data": e.data,
            "created_at": e.created_at,
        }
        for e in events
    ]


@api_router.post("/sync/push")
async def push_sync_events(body: dict[str, object]) -> dict[str, int]:
    """P2-5: Accept pushed events from a remote instance."""
    global _event_store
    if _event_store is None:
        raise HTTPException(status_code=503, detail="Server not initialized")
    count = 0
    raw_events = body.get("events", [])
    if not isinstance(raw_events, list):
        raw_events = []
    for evt_data in raw_events:
        if not isinstance(evt_data, dict):
            continue
        try:
            await _project_events(
                evt_data["aggregate_id"],
                [{"type": evt_data.get("type", ""), "data": evt_data.get("data", {})}],
            )
            count += 1
        except Exception:
            pass
    return {"pushed": count}


_XLSX_TEMPLATE = """import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Example headers
headers = ["Column A", "Column B", "Column C"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Example data
for row in range(2, 11):
    for col in range(1, 4):
        ws.cell(row=row, column=col, value=f"Row{row}Col{col}")

# Auto-fit columns
for col in range(1, 4):
    ws.column_dimensions[get_column_letter(col)].width = 15

output_path = "/tmp/cscode-outputs/output.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
"""

# ---------------------------------------------------------------------------
# P2-1: Register /api/session/... (singular) aliases for all /api/sessions/...
# ---------------------------------------------------------------------------


def _register_session_aliases() -> None:
    """Add singular-path aliases for every /sessions/ route.

    Note: route.path includes the router prefix (/api/...), so we construct
    alias_path relative to the router prefix (strip /api/ before re-adding
    via api_router.add_api_route).
    """
    for route in list(api_router.routes) + list(sessions_router.routes):
        path: str | None = getattr(route, "path", None)
        methods: set[str] | None = getattr(route, "methods", None)
        endpoint = getattr(route, "endpoint", None)
        if not path or not methods or not endpoint:
            continue
        if "/sessions" not in path:
            continue
        # path = /api/sessions/... or /api/sessions → alias = /session/... or /session
        if path.endswith("/api/sessions"):
            alias_path = path.replace("/api/sessions", "/session", 1)
        else:
            alias_path = path.replace("/api/sessions/", "/session/", 1)
        if alias_path == path:
            continue
        for method in methods:
            api_router.add_api_route(
                alias_path,
                endpoint,
                methods=[method],
                include_in_schema=False,
            )


_register_session_aliases()
app.include_router(api_router)

# Static files for web UI — support both development and PyInstaller bundle paths

_web_dist_candidates = [
    Path(__file__).parent.parent / "web" / "dist",  # dev: src/cscode/web/dist/
    Path(__file__).parent.parent.parent / "web" / "dist",  # dev alt: src/web/dist/
]
# PyInstaller: sys._MEIPASS points to the temp extraction root
_mei_path: str | None = getattr(sys, "_MEIPASS", None)
if _mei_path is not None:
    _mei = Path(_mei_path)
    _web_dist_candidates.insert(0, _mei / "web" / "dist")
    _web_dist_candidates.append(_mei.parent / "web" / "dist")

web_dist: Path | None = None
for p in _web_dist_candidates:
    resolved = p.resolve()
    if resolved.is_dir() and (resolved / "index.html").exists():
        web_dist = resolved
        logger.info("Serving static files from: %s", resolved)
        break

if web_dist is not None:
    app.mount("/", StaticFiles(directory=str(web_dist), html=True), name="static")
else:
    logger.warning(
        "No static frontend dist found; tried: %s", [str(p) for p in _web_dist_candidates]
    )
